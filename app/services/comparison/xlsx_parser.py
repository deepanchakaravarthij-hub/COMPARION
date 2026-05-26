from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.cell import get_column_letter  # type: ignore[import-untyped]


@dataclass(frozen=True)
class XlsxCell:
    address: str
    value: str | None
    formula: str | None
    number_format: str
    style_signature: str


@dataclass(frozen=True)
class XlsxSheet:
    name: str
    index: int
    hidden: bool
    metadata: dict[str, Any]
    cells: dict[str, XlsxCell]
    rows_with_data: set[int]
    columns_with_data: set[str]
    hidden_rows: set[int]
    hidden_columns: set[str]


@dataclass(frozen=True)
class XlsxWorkbookModel:
    sheets: list[XlsxSheet]

    @property
    def by_name(self) -> dict[str, XlsxSheet]:
        return {sheet.name: sheet for sheet in self.sheets}


def parse_xlsx(content: bytes) -> XlsxWorkbookModel:
    workbook = load_workbook(filename=BytesIO(content), data_only=False)
    sheets = [
        _parse_sheet(sheet, index) for index, sheet in enumerate(workbook.worksheets, start=1)
    ]
    return XlsxWorkbookModel(sheets=sheets)


def normalize_formula(formula: str | None) -> str | None:
    if not formula:
        return None
    cleaned = "".join(formula.strip().split())
    if not cleaned:
        return None
    if not cleaned.startswith("="):
        cleaned = f"={cleaned}"
    return cleaned.upper()


def _parse_sheet(sheet: Any, index: int) -> XlsxSheet:
    cells: dict[str, XlsxCell] = {}
    rows_with_data: set[int] = set()
    columns_with_data: set[str] = set()

    max_row = max(sheet.max_row, 1)
    max_column = max(sheet.max_column, 1)
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            raw_formula = cell.value if cell.data_type == "f" else None
            normalized_formula = normalize_formula(
                raw_formula if isinstance(raw_formula, str) else None
            )
            normalized_value = _normalize_value(cell.value)
            number_format = cell.number_format or "General"
            style_signature = _style_signature(cell)
            if (
                normalized_value is None
                and normalized_formula is None
                and number_format == "General"
                and style_signature == "style_id:0"
            ):
                continue

            address = cell.coordinate
            cells[address] = XlsxCell(
                address=address,
                value=normalized_value,
                formula=normalized_formula,
                number_format=number_format,
                style_signature=style_signature,
            )
            rows_with_data.add(cell.row)
            columns_with_data.add(get_column_letter(cell.column))

    hidden_rows = {
        index for index, dim in sheet.row_dimensions.items() if getattr(dim, "hidden", False)
    }
    hidden_columns = {
        key for key, dim in sheet.column_dimensions.items() if getattr(dim, "hidden", False)
    }

    metadata = {
        "title": sheet.title,
        "sheet_state": sheet.sheet_state,
        "tab_color": str(sheet.sheet_properties.tabColor.rgb)
        if sheet.sheet_properties.tabColor and sheet.sheet_properties.tabColor.rgb
        else None,
    }
    return XlsxSheet(
        name=sheet.title,
        index=index,
        hidden=sheet.sheet_state != "visible",
        metadata=metadata,
        cells=cells,
        rows_with_data=rows_with_data,
        columns_with_data=columns_with_data,
        hidden_rows=hidden_rows,
        hidden_columns=hidden_columns,
    )


def _normalize_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float):
        return f"{value:.10g}"
    return str(value).strip()


def _style_signature(cell: Any) -> str:
    return (
        f"style_id:{cell.style_id}|font:{cell.font.name}:{cell.font.sz}:{cell.font.bold}:"
        f"{cell.font.italic}:{cell.font.color.index if cell.font.color else None}|"
        f"fill:{cell.fill.fill_type}:{cell.fill.fgColor.index if cell.fill.fgColor else None}|"
        f"alignment:{cell.alignment.horizontal}:{cell.alignment.vertical}:{cell.alignment.wrap_text}|"
        f"number_format:{cell.number_format}"
    )
